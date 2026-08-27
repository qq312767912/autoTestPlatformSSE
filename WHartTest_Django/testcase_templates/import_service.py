"""
用例导入服务

根据导入模版配置，解析 Excel 文件并创建测试用例
"""
import re
import io
import logging
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass, field
from openpyxl import load_workbook
from django.db import transaction

from testcases.models import TestCase, TestCaseStep, TestCaseModule
from projects.models import Project
from .models import ImportExportTemplate

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """导入结果"""
    success: bool = True
    total_rows: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    duplicate_names: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    created_testcases: List[int] = field(default_factory=list)


class TestCaseImportService:
    """测试用例导入服务"""

    MODULE_HIERARCHY_HEADER_PATTERN = re.compile(
        r'^([一二三四五六七八九十]|\d+)级模块$'
    )
    CHINESE_MODULE_LEVELS = {
        '一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
        '六': 6, '七': 7, '八': 8, '九': 9, '十': 10,
    }
    DEFAULT_LEVEL_MAPPING = {
        '最高': 'P0', '高': 'P0', '中': 'P1', '低': 'P2', '最低': 'P3',
    }
    DEFAULT_TEST_TYPE_MAPPING = {
        '冒烟测试': 'smoke', '功能测试': 'functional', '边界测试': 'boundary',
        '异常测试': 'exception', '权限测试': 'permission', '安全测试': 'security',
        '兼容性测试': 'compatibility',
    }

    def __init__(self, template: ImportExportTemplate, project: Project, user):
        self.template = template
        self.project = project
        self.user = user
        self.result = ImportResult()
        self._module_cache: Dict[str, TestCaseModule] = {}
        self._existing_names: Dict[str, List[int]] = {}  # {(module_id, name): [testcase_ids]}

    def import_from_file(self, file) -> ImportResult:
        """从文件导入用例"""
        try:
            workbook = load_workbook(filename=io.BytesIO(file.read()), data_only=True)

            # 选择工作表
            if self.template.sheet_name and self.template.sheet_name in workbook.sheetnames:
                worksheet = workbook[self.template.sheet_name]
            else:
                worksheet = workbook.active

            # 解析合并单元格，将值填充到合并区域内所有单元格
            self._resolve_merged_cells(worksheet)

            # 读取表头
            headers = self._read_headers(worksheet)
            if not headers:
                self.result.success = False
                self.result.errors.append({'row': self.template.header_row, 'error': '无法读取表头'})
                return self.result

            # 预加载现有用例名称（用于检测重复）
            self._load_existing_names()

            # 根据步骤解析模式处理数据
            if self.template.step_parsing_mode == 'multi_row':
                self._import_multi_row_mode(worksheet, headers)
            else:
                self._import_single_cell_mode(worksheet, headers)

            workbook.close()

        except Exception as e:
            logger.exception("导入用例时发生错误")
            self.result.success = False
            self.result.errors.append({'row': 0, 'error': f'导入失败: {str(e)}'})

        return self.result

    def _read_headers(self, worksheet) -> Dict[str, int]:
        """读取表头，返回列名到列索引的映射"""
        headers = {}
        for col_idx, cell in enumerate(worksheet[self.template.header_row]):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        return headers

    @staticmethod
    def _resolve_merged_cells(worksheet):
        """将合并单元格的值填充到合并区域内所有单元格，然后取消合并"""
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            top_left_value = worksheet.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
            worksheet.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=row, column=col).value = top_left_value

    def _load_existing_names(self):
        """预加载项目中现有的用例名称"""
        existing = TestCase.objects.filter(project=self.project).values('id', 'name', 'module_id')
        for tc in existing:
            key = (tc['module_id'], tc['name'])
            if key not in self._existing_names:
                self._existing_names[key] = []
            self._existing_names[key].append(tc['id'])

    def _get_cell_value(self, row, headers: Dict[str, int], excel_column: str) -> str:
        """获取单元格值"""
        if not excel_column or excel_column not in headers:
            return ''
        col_idx = headers[excel_column]
        if col_idx < len(row):
            cell = row[col_idx]
            return str(cell.value).strip() if cell.value is not None else ''
        return ''

    def _get_field_value(self, row, headers: Dict[str, int], field_name: str) -> str:
        """根据字段映射获取并转换字段值"""
        excel_column = self.template.field_mappings.get(field_name)
        if not excel_column:
            return ''

        raw_value = self._get_cell_value(row, headers, excel_column)

        # 应用值转换
        return self.template.transform_value(field_name, raw_value)

    def _get_level(self, row, headers: Dict[str, int]) -> str:
        """读取并标准化用例等级，兼容旧模板中的中文优先级。"""
        value = self._get_field_value(row, headers, 'level').strip()
        if not value:
            return 'P2'
        normalized = value.upper()
        if normalized in {'P0', 'P1', 'P2', 'P3'}:
            return normalized
        return self.DEFAULT_LEVEL_MAPPING.get(value, 'P2')

    def _get_test_type(self, row, headers: Dict[str, int]) -> str:
        """读取并标准化用例类型，兼容中文名称和系统内部代码。"""
        value = self._get_field_value(row, headers, 'test_type').strip()
        if not value:
            return 'functional'
        valid_codes = {choice[0] for choice in TestCase.TEST_TYPE_CHOICES}
        if value in valid_codes:
            return value
        return self.DEFAULT_TEST_TYPE_MAPPING.get(value, 'functional')

    def _get_or_create_module(self, module_path: str) -> Optional[TestCaseModule]:
        """根据模块路径获取或创建模块"""
        if not module_path:
            return None

        # 检查缓存
        if module_path in self._module_cache:
            return self._module_cache[module_path]

        # 解析路径
        delimiter = self.template.module_path_delimiter
        parts = [p.strip() for p in module_path.split(delimiter) if p.strip()]

        if not parts:
            return None

        parent = None
        current_path = ''

        for level, name in enumerate(parts, start=1):
            current_path = f"{current_path}{delimiter}{name}" if current_path else name

            # 先检查缓存
            if current_path in self._module_cache:
                parent = self._module_cache[current_path]
                continue

            # 查找或创建模块
            module, created = TestCaseModule.objects.get_or_create(
                project=self.project,
                parent=parent,
                name=name,
                defaults={
                    'level': level,
                    'creator': self.user,
                }
            )

            self._module_cache[current_path] = module
            parent = module

        return parent

    def _get_module_path(self, row, headers: Dict[str, int]) -> str:
        """根据模板的模块解析模式生成统一的模块路径。"""
        hierarchy_columns = self._get_module_hierarchy_columns(headers)
        if not hierarchy_columns:
            return self._get_field_value(row, headers, 'module')

        values = [self._get_cell_value(row, headers, column) for column in hierarchy_columns]

        # 允许末尾层级为空，但不允许跳过中间层级，否则父子关系不明确。
        first_empty_index = next((index for index, value in enumerate(values) if not value), len(values))
        if any(values[first_empty_index + 1:]):
            raise ValueError('模块层级存在断层，请检查所选表头列及Excel数据。')

        return self.template.module_path_delimiter.join(values[:first_empty_index])

    def _get_module_hierarchy_columns(self, headers: Dict[str, int]) -> List[str]:
        """获取本次导入使用的模块层级列，并兼容旧版单列模块模板。"""
        if self.template.module_parsing_mode == 'columns':
            return self.template.module_hierarchy_columns or []

        # 旧模板通常会把“三级模块”等某一层映射成所属模块。若当前工作表
        # 同时存在多个标准层级表头，则自动使用所有层级列，避免只导入叶子模块。
        mapped_module_column = (self.template.field_mappings or {}).get('module')
        detected_by_level = {}
        for header in headers:
            match = self.MODULE_HIERARCHY_HEADER_PATTERN.fullmatch(header)
            if not match:
                continue
            level_text = match.group(1)
            level = self.CHINESE_MODULE_LEVELS.get(level_text)
            if level is None:
                level = int(level_text)
            if 1 <= level <= 10:
                detected_by_level[level] = header

        # 只采用从一级开始的连续层级。文档只有一至三级时返回3列，
        # 一至六级时返回6列；中间缺级不会被错误地压缩成连续父子关系。
        detected_columns = []
        for level in range(1, 11):
            header = detected_by_level.get(level)
            if not header:
                break
            detected_columns.append(header)
        if mapped_module_column in detected_columns and len(detected_columns) >= 2:
            return detected_columns

        return []

    def _parse_steps_single_cell(self, steps_text: str, expected_text: str) -> List[Dict[str, str]]:
        """
        解析单元格合并模式的步骤

        支持格式:
        - [1]步骤1\n[2]步骤2
        - 1. 步骤1\n2. 步骤2
        - 纯文本（作为单个步骤）
        """
        steps = []

        if not steps_text and not expected_text:
            return steps

        # 尝试匹配 [n]内容 格式
        step_pattern = r'\[(\d+)\]([^\[]*)'
        step_matches = re.findall(step_pattern, steps_text or '')
        expected_matches = re.findall(step_pattern, expected_text or '')

        if step_matches:
            expected_dict = {int(num): content.strip() for num, content in expected_matches}
            # 预期结果无编号时，整段文本赋给最后一个步骤
            fallback_expected = (expected_text or '').strip() if not expected_matches else ''

            for idx, (num, content) in enumerate(step_matches):
                step_num = int(num)
                expected = expected_dict.get(step_num, '')
                if not expected and fallback_expected and idx == len(step_matches) - 1:
                    expected = fallback_expected
                steps.append({
                    'step_number': step_num,
                    'description': content.strip(),
                    'expected_result': expected,
                })
        else:
            # 尝试匹配 n. 内容 格式
            step_pattern2 = r'(\d+)\.\s*([^\d]+?)(?=\d+\.|$)'
            step_matches2 = re.findall(step_pattern2, steps_text or '', re.DOTALL)
            expected_matches2 = re.findall(step_pattern2, expected_text or '', re.DOTALL)

            if step_matches2:
                expected_dict = {int(num): content.strip() for num, content in expected_matches2}
                # 预期结果无编号时，整段文本赋给最后一个步骤
                fallback_expected = (expected_text or '').strip() if not expected_matches2 else ''

                for idx, (num, content) in enumerate(step_matches2):
                    step_num = int(num)
                    expected = expected_dict.get(step_num, '')
                    if not expected and fallback_expected and idx == len(step_matches2) - 1:
                        expected = fallback_expected
                    steps.append({
                        'step_number': step_num,
                        'description': content.strip(),
                        'expected_result': expected,
                    })
            elif steps_text:
                # 按换行分割，每行一个步骤
                lines = [l.strip() for l in steps_text.split('\n') if l.strip()]
                expected_lines = [l.strip() for l in (expected_text or '').split('\n') if l.strip()]

                for i, line in enumerate(lines):
                    steps.append({
                        'step_number': i + 1,
                        'description': line,
                        'expected_result': expected_lines[i] if i < len(expected_lines) else '',
                    })

        return steps

    def _import_single_cell_mode(self, worksheet, headers: Dict[str, int]):
        """单元格合并模式导入"""
        step_config = self.template.step_config or {}
        step_column = step_config.get('step_column') or self.template.field_mappings.get('steps', '')
        expected_column = step_config.get('expected_column') or self.template.field_mappings.get('expected_results', '')

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=self.template.data_start_row),
            start=self.template.data_start_row
        ):
            self.result.total_rows += 1

            try:
                # 获取用例名称
                name = self._get_field_value(row, headers, 'name')
                if not name:
                    # 跳过空行
                    self.result.skipped_count += 1
                    continue

                # 获取模块
                module_path = self._get_module_path(row, headers)
                module = self._get_or_create_module(module_path)

                if not module:
                    self.result.errors.append({
                        'row': row_idx,
                        'error': f'无法解析模块路径: {module_path}',
                        'name': name,
                    })
                    self.result.error_count += 1
                    continue

                # 检查重复
                key = (module.id, name)
                if key in self._existing_names:
                    self.result.duplicate_names.append({
                        'row': row_idx,
                        'name': name,
                        'module': module_path,
                        'existing_ids': self._existing_names[key],
                    })

                # 获取其他字段
                level = self._get_level(row, headers)
                test_type = self._get_test_type(row, headers)
                precondition = self._get_field_value(row, headers, 'precondition')
                notes = self._get_field_value(row, headers, 'notes')

                # 获取步骤
                steps_text = self._get_cell_value(row, headers, step_column)
                expected_text = self._get_cell_value(row, headers, expected_column)
                steps = self._parse_steps_single_cell(steps_text, expected_text)

                # 创建用例
                with transaction.atomic():
                    testcase = TestCase.objects.create(
                        project=self.project,
                        module=module,
                        name=name,
                        level=level,
                        test_type=test_type,
                        precondition=precondition,
                        notes=notes,
                        creator=self.user,
                    )

                    # 创建步骤
                    for step_data in steps:
                        TestCaseStep.objects.create(
                            test_case=testcase,
                            step_number=step_data['step_number'],
                            description=step_data['description'],
                            expected_result=step_data['expected_result'],
                            creator=self.user,
                        )

                    self.result.imported_count += 1
                    self.result.created_testcases.append(testcase.id)

                    # 更新现有名称缓存
                    if key not in self._existing_names:
                        self._existing_names[key] = []
                    self._existing_names[key].append(testcase.id)

            except Exception as e:
                logger.exception(f"导入第 {row_idx} 行时发生错误")
                self.result.errors.append({
                    'row': row_idx,
                    'error': str(e),
                })
                self.result.error_count += 1

    def _import_multi_row_mode(self, worksheet, headers: Dict[str, int]):
        """多行步骤模式导入"""
        step_config = self.template.step_config or {}
        case_identifier_column = step_config.get('case_identifier_column') or self.template.field_mappings.get('name', '')
        step_column = step_config.get('step_column') or self.template.field_mappings.get('steps', '')
        expected_column = step_config.get('expected_column') or self.template.field_mappings.get('expected_results', '')

        # 收集用例数据
        current_case = None
        cases_to_create = []

        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=self.template.data_start_row),
            start=self.template.data_start_row
        ):
            self.result.total_rows += 1

            # 获取用例标识列的值
            case_name = self._get_cell_value(row, headers, case_identifier_column)

            if case_name:
                # 新用例开始
                if current_case:
                    cases_to_create.append(current_case)

                current_case = {
                    'row': row_idx,
                    'name': case_name,
                    'module_path': self._get_module_path(row, headers),
                    'level': self._get_level(row, headers),
                    'test_type': self._get_test_type(row, headers),
                    'precondition': self._get_field_value(row, headers, 'precondition'),
                    'notes': self._get_field_value(row, headers, 'notes'),
                    'steps': [],
                }

            # 添加步骤
            if current_case:
                step_desc = self._get_cell_value(row, headers, step_column)
                expected = self._get_cell_value(row, headers, expected_column)

                if step_desc or expected:
                    current_case['steps'].append({
                        'step_number': len(current_case['steps']) + 1,
                        'description': step_desc,
                        'expected_result': expected,
                    })

        # 添加最后一个用例
        if current_case:
            cases_to_create.append(current_case)

        # 批量创建用例
        for case_data in cases_to_create:
            try:
                module = self._get_or_create_module(case_data['module_path'])
                if not module:
                    self.result.errors.append({
                        'row': case_data['row'],
                        'error': f'无法解析模块路径: {case_data["module_path"]}',
                        'name': case_data['name'],
                    })
                    self.result.error_count += 1
                    continue

                # 检查重复
                key = (module.id, case_data['name'])
                if key in self._existing_names:
                    self.result.duplicate_names.append({
                        'row': case_data['row'],
                        'name': case_data['name'],
                        'module': case_data['module_path'],
                        'existing_ids': self._existing_names[key],
                    })

                with transaction.atomic():
                    testcase = TestCase.objects.create(
                        project=self.project,
                        module=module,
                        name=case_data['name'],
                        level=case_data['level'],
                        test_type=case_data['test_type'],
                        precondition=case_data['precondition'],
                        notes=case_data['notes'],
                        creator=self.user,
                    )

                    for step_data in case_data['steps']:
                        TestCaseStep.objects.create(
                            test_case=testcase,
                            step_number=step_data['step_number'],
                            description=step_data['description'],
                            expected_result=step_data['expected_result'],
                            creator=self.user,
                        )

                    self.result.imported_count += 1
                    self.result.created_testcases.append(testcase.id)

                    if key not in self._existing_names:
                        self._existing_names[key] = []
                    self._existing_names[key].append(testcase.id)

            except Exception as e:
                logger.exception(f"导入用例 {case_data['name']} 时发生错误")
                self.result.errors.append({
                    'row': case_data['row'],
                    'error': str(e),
                    'name': case_data['name'],
                })
                self.result.error_count += 1
