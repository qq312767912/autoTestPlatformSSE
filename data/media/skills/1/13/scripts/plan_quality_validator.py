# -*- coding: utf-8 -*-
"""执行AI生成结构自检，或在人工评审完成后计算八项指标。

用法：
  python plan_quality_validator.py \
    --requirements 原子需求清单.xlsx \
    --plan test_plan.xlsx \
    --conflicts 冲突清单.xlsx \
    --mode self-check --out 生成结构自检.xlsx

人工评审完成后使用 `--mode reviewed` 计算八项指标。AI生成阶段不得使用 reviewed 模式。

测试方案明细建议列：方案编号、测试场景类型、测试要点、原子需求ID、来源定位、
来源类型、确认状态、人工评审状态。人工评审状态支持：accepted、minor_edit、
major_edit、deleted、duplicate；人工新增行的来源类型使用 human_add。
原子需求人工标注字段：coverage_review_status=covered/uncovered，
completeness_review_status=complete/incomplete/not_applicable。
"""

import argparse
import json
import re
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook


ALIASES = {
    "point_id": ["方案编号", "测试点ID", "测试点编号"],
    "scenario": ["测试场景类型", "场景类型"],
    "point": ["测试要点", "测试点", "验证内容"],
    "requirement_ids": ["原子需求ID", "关联需求ID", "需求ID"],
    "source": ["来源定位", "需求来源", "来源位置"],
    "source_type": ["来源类型", "生成类型"],
    "confirm_status": ["确认状态"],
    "review_status": ["人工评审状态", "评审状态"],
}


def load_json(path, default):
    if not path:
        return default
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def row_dicts(path, sheet_name):
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"{path} 缺少工作表：{sheet_name}")
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def load_requirements(path):
    if str(path).lower().endswith(".json"):
        return records(load_json(path, []))
    items = []
    for row in row_dicts(path, "原子需求清单"):
        items.append({
            "requirement_id": row.get("需求ID") or row.get("原子需求ID"),
            "priority": row.get("优先级"),
            "status": row.get("确认状态") or "confirmed",
            "coverage_review_status": row.get("覆盖评审"),
            "completeness_review_status": row.get("完整性评审"),
        })
    return items


def load_conflicts(path):
    if not path:
        return []
    if str(path).lower().endswith(".json"):
        return records(load_json(path, []))
    return [
        {"status": row.get("处理状态") or "unresolved", "id": row.get("冲突ID")}
        for row in row_dicts(path, "冲突清单")
    ]


def records(value):
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        for key in ("requirements", "items", "conflicts"):
            if isinstance(value.get(key), list):
                return value[key]
    return []


def find_columns(headers):
    result = {}
    normalized = [str(item or "").strip() for item in headers]
    for field, aliases in ALIASES.items():
        result[field] = next((normalized.index(alias) for alias in aliases if alias in normalized), None)
    return result


def cell(row, index):
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def split_ids(value):
    return [item.strip() for item in re.split(r"[,，;；\s]+", value) if item.strip()]


def normalize_point(value):
    value = re.sub(r"[\s，。；、,.!！?？:：()（）\[\]【】]", "", value.lower())
    return re.sub(r"^(验证|检查|确认)", "", value)


def rate(numerator, denominator):
    if not denominator:
        return None
    return round(numerator / denominator, 4)


def metric(numerator=None, denominator=None, reason=""):
    if numerator is None or denominator is None:
        return {"value": None, "display": "N/A", "reason": reason}
    value = rate(numerator, denominator)
    return {
        "numerator": numerator,
        "denominator": denominator,
        "value": value,
        "display": f"{value:.1%}" if value is not None else "N/A",
    }


def save_xlsx(result, output_path):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "结构自检" if result["mode"] == "self-check" else "指标汇总"
    summary.append(["项目", "结果", "说明"])
    summary.append(["模式", result["mode"], result["notice"]])
    count_labels = {
        "atomic_requirements": "原子需求数",
        "ai_test_points": "AI测试点数",
        "human_added_points": "人工新增测试点数",
        "requirements_with_test_point_links": "已建立测试点关联的需求数",
        "duplicates": "疑似重复测试点数",
        "unresolved_conflicts": "未解决冲突数",
    }
    for key, value in result["counts"].items():
        summary.append([count_labels.get(key, key), value, ""])
    if result.get("metrics"):
        summary.append([])
        summary.append(["指标", "结果", "计算口径"])
        for name, value in result["metrics"].items():
            formula = ""
            if value.get("numerator") is not None:
                formula = f"{value['numerator']} / {value['denominator']}"
            summary.append([name, value.get("display"), formula or value.get("reason", "")])

    issues = workbook.create_sheet("问题明细")
    issues.append(["类型", "对象", "说明"])
    for item in result["structural_self_check"]["issues"]:
        issues.append(["结构问题", "", item])
    for item in result.get("review_warnings", []):
        issues.append(["评审数据缺口", "", item])
    for requirement_id in result["details"]["uncovered_requirement_ids"]:
        issues.append(["无测试点关联", requirement_id, "需人工判断是否真实遗漏"])
    for duplicate in result["details"]["duplicate_rows"]:
        issues.append(["疑似重复", f"行{duplicate['row']}", f"疑似与行{duplicate['duplicate_of']}重复"])
    for column in result["details"]["missing_columns"]:
        issues.append(["缺少字段", column, "请补充后重新检查"])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            header_font = copy(cell.font)
            header_font.bold = True
            cell.font = header_font
        for column_cells in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--requirements", required=True)
    parser.add_argument("--plan", required=True)
    parser.add_argument("--sheet", default="测试方案明细")
    parser.add_argument("--conflicts")
    parser.add_argument("--out")
    parser.add_argument("--mode", choices=["self-check", "reviewed"], default="self-check")
    parser.add_argument("--duplicate-threshold", type=float, default=0.90)
    args = parser.parse_args()

    requirements = load_requirements(args.requirements)
    requirement_map = {
        str(item.get("requirement_id", "")).strip(): item
        for item in requirements
        if item.get("requirement_id")
        and item.get("status", "confirmed") not in {"obsolete", "rejected"}
    }

    workbook = load_workbook(args.plan, data_only=True, read_only=True)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"缺少工作表：{args.sheet}")
    rows = list(workbook[args.sheet].iter_rows(values_only=True))
    if not rows:
        raise SystemExit("测试方案明细为空")
    columns = find_columns(rows[0])
    missing = [name for name in ("point", "requirement_ids", "source") if columns[name] is None]

    plan_rows = []
    for row_number, row in enumerate(rows[1:], start=2):
        point = cell(row, columns["point"])
        if not point:
            continue
        plan_rows.append({
            "row": row_number,
            "point_id": cell(row, columns["point_id"]),
            "point": point,
            "scenario": cell(row, columns["scenario"]).lower(),
            "requirement_ids": split_ids(cell(row, columns["requirement_ids"])),
            "source": cell(row, columns["source"]),
            "source_type": cell(row, columns["source_type"]).lower() or "requirement",
            "confirm_status": cell(row, columns["confirm_status"]).lower(),
            "review_status": cell(row, columns["review_status"]).lower(),
        })

    ai_rows = [row for row in plan_rows if row["source_type"] != "human_add"]
    active_rows = [row for row in plan_rows if row["review_status"] != "deleted"]
    covered_ids = {
        req_id for row in active_rows for req_id in row["requirement_ids"] if req_id in requirement_map
    }
    traceable = [
        row for row in ai_rows
        if row["source"] and row["requirement_ids"]
        and all(req_id in requirement_map for req_id in row["requirement_ids"])
    ]

    duplicate_rows = []
    normalized = []
    for row in ai_rows:
        text = normalize_point(row["point"])
        duplicate_of = None
        for previous, previous_text in normalized:
            if text and SequenceMatcher(None, text, previous_text).ratio() >= args.duplicate_threshold:
                duplicate_of = previous
                break
        if duplicate_of:
            duplicate_rows.append({"row": row["row"], "duplicate_of": duplicate_of["row"]})
        else:
            normalized.append((row, text))

    scenarios_by_requirement = {}
    for row in active_rows:
        for req_id in row["requirement_ids"]:
            scenarios_by_requirement.setdefault(req_id, set()).add(row["scenario"])
    multi_requirements = []
    complete_multi = 0
    for req_id, requirement in requirement_map.items():
        required = requirement.get("required_scenarios") or []
        if requirement.get("needs_multiple_scenarios") and not required:
            required = ["正向", "反向", "边界"]
        if required:
            multi_requirements.append(req_id)
            actual = scenarios_by_requirement.get(req_id, set())
            if all(any(expected.lower() in value for value in actual) for expected in required):
                complete_multi += 1

    reviewed = [row for row in ai_rows if row["review_status"]]
    accurate = [row for row in reviewed if row["review_status"] in {"accepted", "minor_edit"}]
    accepted = [row for row in reviewed if row["review_status"] == "accepted"]
    deleted = [row for row in reviewed if row["review_status"] == "deleted"]
    human_added = [row for row in plan_rows if row["source_type"] == "human_add"]

    conflicts = load_conflicts(args.conflicts)
    unresolved = [item for item in conflicts if item.get("status", "unresolved") not in {"resolved", "accepted"}]

    structural_issues = []
    if set(requirement_map) - covered_ids:
        structural_issues.append("存在未关联测试点的原子需求；需人工判断是否真实遗漏")
    if len(traceable) != len(ai_rows):
        structural_issues.append("存在缺少需求ID或来源定位的测试点")
    if duplicate_rows:
        structural_issues.append("存在文本高度相似的疑似重复点，需人工确认")
    if unresolved:
        structural_issues.append(f"存在{len(unresolved)}条未解决业务冲突")
    if missing:
        structural_issues.append(f"方案缺少结构字段：{', '.join(missing)}")

    metrics = None
    review_warnings = []
    if args.mode == "reviewed":
        coverage_labels = [item.get("coverage_review_status") for item in requirement_map.values()]
        completeness_labels = [
            item.get("completeness_review_status") for item in requirement_map.values()
            if item.get("completeness_review_status") != "not_applicable"
        ]
        all_points_reviewed = len(reviewed) == len(ai_rows)
        all_requirements_reviewed = bool(requirement_map) and all(
            value in {"covered", "uncovered"} for value in coverage_labels
        )
        all_completeness_reviewed = bool(completeness_labels) and all(
            value in {"complete", "incomplete"} for value in completeness_labels
        )
        if not all_points_reviewed:
            review_warnings.append("并非所有AI测试点都已完成人工评审")
        if not all_requirements_reviewed:
            review_warnings.append("并非所有原子需求都已标注covered/uncovered")
        if not all_completeness_reviewed:
            review_warnings.append("需要多场景验证的需求尚未全部标注complete/incomplete")

        reviewed_duplicates = [row for row in ai_rows if row["review_status"] == "duplicate"]
        covered_by_human = sum(value == "covered" for value in coverage_labels)
        complete_by_human = sum(value == "complete" for value in completeness_labels)
        metrics = {
            "需求覆盖率": metric(covered_by_human, len(requirement_map)) if all_requirements_reviewed else metric(reason="人工覆盖评审未完成"),
            "测试点准确率": metric(len(accurate), len(ai_rows)) if all_points_reviewed else metric(reason="人工测试点评审未完成"),
            "测试点完整率": metric(complete_by_human, len(completeness_labels)) if all_completeness_reviewed else metric(reason="人工完整性评审未完成"),
            "测试点可追溯率": metric(len(traceable), len(ai_rows)),
            "人工直接接收率": metric(len(accepted), len(ai_rows)) if all_points_reviewed else metric(reason="人工测试点评审未完成"),
            "人工补充率": metric(len(human_added), len(active_rows)),
            "无效测试点率": metric(len(deleted), len(ai_rows)) if all_points_reviewed else metric(reason="人工测试点评审未完成"),
            "测试点重复率": metric(len(reviewed_duplicates), len(ai_rows)) if all_points_reviewed else metric(reason="人工测试点评审未完成"),
        }

    result = {
        "mode": args.mode,
        "notice": "AI结构自检，不能作为八项指标结果" if args.mode == "self-check" else "基于人工评审标注计算",
        "input": {"requirements": args.requirements, "plan": args.plan, "sheet": args.sheet},
        "counts": {
            "atomic_requirements": len(requirement_map),
            "ai_test_points": len(ai_rows),
            "human_added_points": len(human_added),
            "requirements_with_test_point_links": len(covered_ids),
            "duplicates": len(duplicate_rows),
            "unresolved_conflicts": len(unresolved),
        },
        "metrics": metrics,
        "structural_self_check": {"passed": not structural_issues, "issues": structural_issues},
        "review_warnings": review_warnings,
        "details": {
            "uncovered_requirement_ids": sorted(set(requirement_map) - covered_ids),
            "duplicate_rows": duplicate_rows,
            "missing_columns": missing,
        },
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    if args.out:
        if str(args.out).lower().endswith(".xlsx"):
            save_xlsx(result, args.out)
        else:
            Path(args.out).parent.mkdir(parents=True, exist_ok=True)
            Path(args.out).write_text(output, encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
