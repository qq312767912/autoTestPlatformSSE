#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""统一美化测试用例质量审查 Excel 报告。

模型先生成包含“执行摘要、问题明细、改写示例、待确认事项、治理建议”
的工作簿，再调用本脚本固化字体、配色、列宽、打印和浏览体验。
"""

import argparse
import os
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "16324F"
BLUE = "2563EB"
LIGHT_BLUE = "EAF2FF"
SLATE = "475569"
MUTED = "64748B"
GRID = "D8E1EC"
WHITE = "FFFFFF"
RED = "DC2626"
RED_BG = "FDECEC"
AMBER = "D97706"
AMBER_BG = "FFF4DB"
GREEN = "16834A"
GREEN_BG = "EAF7EF"
SOFT_BG = "F5F7FA"
FONT_NAME = "Microsoft YaHei"

THIN_BORDER = Border(
    left=Side(style="thin", color=GRID),
    right=Side(style="thin", color=GRID),
    top=Side(style="thin", color=GRID),
    bottom=Side(style="thin", color=GRID),
)


def text(value):
    return str(value or "").strip()


def merged_row(sheet, row):
    return any(rng.min_row == row and rng.max_row == row for rng in sheet.merged_cells.ranges)


def merge_across(sheet, row, start, end):
    if end <= start or merged_row(sheet, row):
        return
    sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)


def style_title(sheet, title_row=1, subtitle_row=None):
    merge_across(sheet, title_row, 1, sheet.max_column)
    cell = sheet.cell(title_row, 1)
    cell.font = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[title_row].height = 38
    if subtitle_row and text(sheet.cell(subtitle_row, 1).value):
        merge_across(sheet, subtitle_row, 1, sheet.max_column)
        cell = sheet.cell(subtitle_row, 1)
        cell.font = Font(name=FONT_NAME, size=10, color=MUTED)
        cell.fill = PatternFill("solid", fgColor=SOFT_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[subtitle_row].height = 24


def style_table_header(sheet, row, start=1, end=None):
    end = end or sheet.max_column
    for col in range(start, end + 1):
        cell = sheet.cell(row, col)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    sheet.row_dimensions[row].height = 28


def style_body(sheet, start_row, end_row=None):
    end_row = end_row or sheet.max_row
    for row in range(start_row, end_row + 1):
        if not any(text(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)):
            continue
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else "F8FAFC")
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.font = Font(name=FONT_NAME, size=10, color=NAVY)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="top",
                wrap_text=True,
            )
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 15, 32)


def common_sheet_setup(sheet, landscape=False, repeat_rows=None):
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.page_setup.orientation = "landscape" if landscape else "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.45
    sheet.page_margins.bottom = 0.45
    sheet.oddFooter.center.text = "第 &P 页 / 共 &N 页"
    sheet.oddFooter.center.size = 9
    sheet.oddFooter.center.color = MUTED
    if repeat_rows:
        sheet.print_title_rows = repeat_rows
    sheet.print_area = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def style_issue_sheet(sheet):
    style_title(sheet, 1, 2)
    section_rows = []
    header_rows = []
    for row in range(3, sheet.max_row + 1):
        first = text(sheet.cell(row, 1).value)
        if first.startswith("▎") or (first and all(not text(sheet.cell(row, c).value) for c in range(2, sheet.max_column + 1))):
            section_rows.append(row)
        if first == "序号" and text(sheet.cell(row, 2).value):
            header_rows.append(row)

    for row in section_rows:
        merge_across(sheet, row, 1, sheet.max_column)
        cell = sheet.cell(row, 1)
        label = text(cell.value)
        color = RED if "高" in label else AMBER if "中" in label else GREEN
        bg = RED_BG if "高" in label else AMBER_BG if "中" in label else GREEN_BG
        cell.font = Font(name=FONT_NAME, size=12, bold=True, color=color)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[row].height = 27

    for index, row in enumerate(header_rows):
        style_table_header(sheet, row)
        end = (section_rows[index + 1] - 1) if index + 1 < len(section_rows) else sheet.max_row
        style_body(sheet, row + 1, end)

    widths = [8, 30, 29, 12, 20, 32, 40]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "B6" if header_rows else "A3"
    if header_rows:
        sheet.auto_filter.ref = f"A{header_rows[0]}:G{sheet.max_row}"
    common_sheet_setup(sheet, landscape=True, repeat_rows=f"{header_rows[0]}:{header_rows[0]}" if header_rows else None)


def style_summary_sheet(sheet):
    style_title(sheet, 1)
    for row in range(2, min(sheet.max_row, 4) + 1):
        sheet.cell(row, 1).font = Font(name=FONT_NAME, size=10, bold=True, color=SLATE)
        sheet.cell(row, 2).font = Font(name=FONT_NAME, size=11, bold=True, color=BLUE)
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).fill = PatternFill("solid", fgColor=SOFT_BG)
            sheet.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 25

    for row in range(2, sheet.max_row + 1):
        first = text(sheet.cell(row, 1).value)
        if first in {"严重程度分布", "最优先处理的风险"}:
            merge_across(sheet, row, 1, sheet.max_column)
            cell = sheet.cell(row, 1)
            cell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[row].height = 28

    distribution_header = next((r for r in range(2, sheet.max_row + 1) if text(sheet.cell(r, 1).value) == "严重程度分布"), None)
    if distribution_header:
        style_table_header(sheet, distribution_header)
        start = distribution_header + 1
        end = start
        while end <= sheet.max_row and text(sheet.cell(end, 1).value):
            end += 1
        style_body(sheet, start, end - 1)
        colors = {"高": (RED, RED_BG), "中": (AMBER, AMBER_BG), "低": (GREEN, GREEN_BG)}
        for row in range(start, end):
            label = text(sheet.cell(row, 1).value)
            for key, (fg, bg) in colors.items():
                if key in label:
                    sheet.cell(row, 1).font = Font(name=FONT_NAME, size=10, bold=True, color=fg)
                    sheet.cell(row, 1).fill = PatternFill("solid", fgColor=bg)
                    break

    for row in range(2, sheet.max_row + 1):
        if text(sheet.cell(row, 1).value).startswith(tuple(f"{i}." for i in range(1, 10))):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                cell.font = Font(name=FONT_NAME, size=10, color=NAVY, bold=(col == 1))
                cell.fill = PatternFill("solid", fgColor="F8FAFC" if row % 2 else WHITE)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="thin", color=GRID))
            sheet.row_dimensions[row].height = 31

    widths = [38, 18, 14, 54]
    for col, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"
    common_sheet_setup(sheet, landscape=True)


def style_standard_sheet(sheet, widths):
    style_title(sheet, 1)
    header_row = next(
        (row for row in range(2, min(sheet.max_row, 8) + 1) if sum(bool(text(sheet.cell(row, c).value)) for c in range(1, sheet.max_column + 1)) >= 2),
        2,
    )
    style_table_header(sheet, header_row)
    style_body(sheet, header_row + 1)
    for col, width in enumerate(widths, start=1):
        if col <= sheet.max_column:
            sheet.column_dimensions[get_column_letter(col)].width = width
    for row in range(header_row + 1, sheet.max_row + 1):
        longest = max((len(text(sheet.cell(row, c).value)) for c in range(1, sheet.max_column + 1)), default=0)
        sheet.row_dimensions[row].height = min(max(32, 18 + longest // 4), 110)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    common_sheet_setup(sheet, landscape=sheet.max_column >= 3, repeat_rows=f"{header_row}:{header_row}")


def add_severity_formatting(sheet):
    severity_col = None
    header_row = None
    for row in range(1, min(sheet.max_row, 10) + 1):
        for col in range(1, sheet.max_column + 1):
            if text(sheet.cell(row, col).value) == "严重程度":
                severity_col, header_row = col, row
                break
    if not severity_col:
        return
    letter = get_column_letter(severity_col)
    target = f"{letter}{header_row + 1}:{letter}{sheet.max_row}"
    for value, fg, bg in (("高", RED, RED_BG), ("中", AMBER, AMBER_BG), ("低", GREEN, GREEN_BG)):
        sheet.conditional_formatting.add(
            target,
            FormulaRule(
                formula=[f'{letter}{header_row + 1}="{value}"'],
                font=Font(name=FONT_NAME, bold=True, color=fg),
                fill=PatternFill("solid", fgColor=bg),
            ),
        )


def beautify(input_path, output_path):
    workbook = load_workbook(input_path)
    preferred = ["执行摘要", "问题明细", "改写示例", "待确认事项", "治理建议"]
    ordered = [workbook[name] for name in preferred if name in workbook.sheetnames]
    ordered.extend(sheet for sheet in workbook.worksheets if sheet not in ordered)
    workbook._sheets = ordered

    for sheet in workbook.worksheets:
        if sheet.title == "问题明细":
            style_issue_sheet(sheet)
            add_severity_formatting(sheet)
        elif sheet.title == "执行摘要":
            style_summary_sheet(sheet)
        elif sheet.title == "改写示例":
            style_standard_sheet(sheet, [18, 52, 70])
        elif sheet.title == "待确认事项":
            style_standard_sheet(sheet, [9, 52, 36, 20])
        elif sheet.title == "治理建议":
            style_standard_sheet(sheet, [22, 90])
        else:
            style_standard_sheet(sheet, [18] + [36] * max(sheet.max_column - 1, 0))

    workbook.calculation.fullCalcOnLoad = True
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="美化测试用例质量审查 Excel 报告")
    parser.add_argument("--input", required=True, help="模型生成的原始 xlsx")
    parser.add_argument("--output", help="美化后的 xlsx；默认输出到 SKILL_OUTPUT_DIR")
    args = parser.parse_args()

    input_path = os.path.abspath(args.input)
    if not os.path.isfile(input_path):
        raise SystemExit(f"输入文件不存在: {args.input}")
    output_path = args.output
    if not output_path:
        output_dir = os.environ.get("SKILL_OUTPUT_DIR") or os.path.dirname(input_path)
        stem = Path(input_path).stem
        output_path = os.path.join(output_dir, f"{stem}_美化版.xlsx")
    output_path = os.path.abspath(output_path)
    beautify(input_path, output_path)
    print(output_path)


if __name__ == "__main__":
    main()
