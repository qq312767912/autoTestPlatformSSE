import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).with_name("beautify_quality_review_report.py")
MODULE_SPEC = importlib.util.spec_from_file_location("beautify_report", MODULE_PATH)
beautify_report = importlib.util.module_from_spec(MODULE_SPEC)
assert MODULE_SPEC and MODULE_SPEC.loader
MODULE_SPEC.loader.exec_module(beautify_report)


class BeautifyQualityReviewReportTests(unittest.TestCase):
    def test_beautify_orders_and_styles_standard_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            output = Path(temp_dir) / "output.xlsx"
            workbook = Workbook()
            issues = workbook.active
            issues.title = "问题明细"
            issues.append(["测试用例质量审查报告"])
            issues.append(["源文件：测试用例.xlsx"])
            issues.append([])
            issues.append(["▎高风险问题"])
            issues.append(["序号", "用例名称", "原文/位置", "严重程度", "问题类型", "问题说明", "修改建议"])
            issues.append([1, "登录", "预期：成功", "高", "模糊措辞", "不可验收", "补充页面状态"])
            summary = workbook.create_sheet("执行摘要")
            summary.append(["执行摘要"])
            summary.append(["用例总数", "10"])
            summary.append([])
            summary.append(["严重程度分布", "数量", "占比", "说明"])
            summary.append(["高风险", "1", "10%", "核心风险"])
            workbook.save(source)

            beautify_report.beautify(source, output)
            result = load_workbook(output)

            self.assertEqual(result.sheetnames[:2], ["执行摘要", "问题明细"])
            self.assertEqual(result["执行摘要"]["A1"].fill.fgColor.rgb, "0016324F")
            self.assertEqual(result["问题明细"]["A5"].fill.fgColor.rgb, "002563EB")
            self.assertEqual(result["问题明细"].freeze_panes, "B6")
            self.assertEqual(result["问题明细"].page_setup.orientation, "landscape")
            self.assertEqual(result["问题明细"].sheet_view.showGridLines, False)


if __name__ == "__main__":
    unittest.main()
